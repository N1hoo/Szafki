# ui/main_window.py
from PyQt6 import QtCore, QtGui, QtWidgets
from PyQt6.QtWidgets import QMainWindow, QDialog, QMessageBox
from PyQt6.QtCore import Qt

from ui_generated.ui_mainwindow import Ui_MainWindow

from ui.dialogs.nowi_pracownicy_dialog import NowiPracownicyDialog
from ui.renderers import render_lockers_table, render_employees_table, render_employee_lockers_panel
from ui.dialogs.login_dialog import LoginDialog, ChangePasswordDialog
from session import get_current_user, clear_current_user
from ui.admin.users_panel import UsersPanel

from ui.handlers.lockers_handlers import LockersHandlersMixin
from ui.handlers.employees_handlers import EmployeesHandlersMixin

from services.lockers_service import LockersService
from services.employees_service import EmployeesService
from services.assignment_service import AssignmentService

from pracownicy import DF_PRACOWNICY

from ui_generated.ui_dodaj_szafki import Ui_DodajSzafki
from ui_generated.ui_dodaj_pracownika import Ui_Dialog

from database import dodaj_szafke, edytuj_szafke, usun_szafke

class OknoGlowne(QMainWindow, Ui_MainWindow, LockersHandlersMixin, EmployeesHandlersMixin):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setupUi(self)

        # Ukrycie starych widgetów filtrów
        try:
            self.groupBox.hide()
        except Exception:
            pass
        try:
            self.MiejsceCB.hide()
        except Exception:
            pass
        try:
            self.label_6.hide()
        except Exception:
            pass

        # Ukrycie starych filtrów pracowników
        try:
            self.groupBox_2.hide()
        except Exception:
            pass

        # Inicjalizacja serwisów
        self.lockers = LockersService()
        self.employees = EmployeesService(self.lockers)
        self.assignment = AssignmentService()

        # Filtry kolumn dla szafek i pracowników
        self._column_filters: dict[int, str] = {}
        self._emp_column_filters: dict[int, str] = {}

        # Domyślnie tylko zatrudnieni
        self._emp_only_active = True
        
        # Domyślnie ukryj zwolnionych
        self._show_dismissed_lockers = False

        # Cache
        self._wszystkie_szafki = []
        self._szafki_by_kod = {}

        # Filtry kolumn (indeks -> tekst filtra)
        self._column_filters: dict[int, str] = {}

        # Dodatkowe elementy UI
        self._init_panel_szafek_pracownika()
        self._init_nowi_pracownicy_button()

        # Responsywne layouty
        self._apply_responsive_layouts()

        # Tekst przycisku zwalniania
        self.UsunPr.setText("Zwolnij wszystkie szafki pracownika")

        self._connect_signals()

        # Menu konta
        self._init_account_menu()
        self._init_export_menu()

        # Zakładki admina (tworzone leniwie)
        self._admin_tabs_created = False

        # Start
        self.reload_all()
        self._update_user_status()
        self._ensure_admin_tabs()

    # ---------- INICJALIZACJA UI ----------
    def _init_nowi_pracownicy_button(self):
        self.NowiPracownicyBt = QtWidgets.QPushButton(self.PracownicyTab)
        self.NowiPracownicyBt.setText("Nowi pracownicy")
        self.NowiPracownicyBt.clicked.connect(lambda: self.sprawdz_nowych_pracownikow(pokaz_zawsze=True))

    def _init_panel_szafek_pracownika(self):
        self.PanelSzafekPr = QtWidgets.QGroupBox(self.PracownicyTab)
        self.PanelSzafekPr.setTitle("Szafki pracownika")
        # Stała szerokość panelu
        try:
            self.PanelSzafekPr.setFixedWidth(480)
        except Exception:
            pass
        lay = QtWidgets.QVBoxLayout(self.PanelSzafekPr)

        self.LblPracownikInfo = QtWidgets.QLabel("Wybierz pracownika z tabeli po lewej.")
        self.LblPracownikInfo.setWordWrap(True)
        lay.addWidget(self.LblPracownikInfo)

        self.TabelaSzafekPr = QtWidgets.QTableWidget(self.PanelSzafekPr)
        self.TabelaSzafekPr.setColumnCount(4)
        self.TabelaSzafekPr.setHorizontalHeaderLabels(["Miejsce", "Nr szafki", "Nr zamka", "ID"])
        self.TabelaSzafekPr.setColumnHidden(3, True)
        self.TabelaSzafekPr.setEditTriggers(QtWidgets.QAbstractItemView.EditTrigger.NoEditTriggers)
        self.TabelaSzafekPr.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectionBehavior.SelectRows)
        self.TabelaSzafekPr.setSelectionMode(QtWidgets.QAbstractItemView.SelectionMode.SingleSelection)
        lay.addWidget(self.TabelaSzafekPr, 1)

        btn_row = QtWidgets.QHBoxLayout()
        self.BtnZwolnijWybranaSz = QtWidgets.QPushButton("Zwolnij zaznaczoną szafkę")
        self.BtnSkoczDoSzafki = QtWidgets.QPushButton("Pokaż w zakładce Szafki")
        btn_row.addWidget(self.BtnZwolnijWybranaSz)
        btn_row.addWidget(self.BtnSkoczDoSzafki)
        lay.addLayout(btn_row)

        self.BtnZwolnijWybranaSz.clicked.connect(self._zwolnij_szafke_z_panelu)
        self.BtnSkoczDoSzafki.clicked.connect(self._skocz_do_szafki_w_zakladce)

    def _apply_responsive_layouts(self):
        # Centralny widget: TabWidget na całą przestrzeń "To ja i moja przestrzeń"
        cw = self.centralwidget
        if cw.layout() is None:
            lay = QtWidgets.QVBoxLayout(cw)
            lay.setContentsMargins(0, 0, 0, 0)
            lay.addWidget(self.SzafkiWidget)

        self._layout_szafki_tab()
        self._layout_pracownicy_tab()

    def _layout_szafki_tab(self):
        tab = self.SzafkiTab
        if tab.layout() is not None:
            return
        root = QtWidgets.QVBoxLayout(tab)
        root.setContentsMargins(10, 10, 10, 10)
        root.setSpacing(8)

        top = QtWidgets.QHBoxLayout()
        top.setSpacing(8)

        top.addWidget(self.groupBox)
        top.addStretch(1)

        btns = QtWidgets.QHBoxLayout()
        btns.setSpacing(8)
        btns.addWidget(self.PrzydzielSz)
        btns.addWidget(self.EdytujSz)
        btns.addWidget(self.ZwolnijSz)
        
        # Dodaj szafki button
        self.DodajSzBt = QtWidgets.QPushButton(tab)
        self.DodajSzBt.setText("Dodaj szafki")
        self.DodajSzBt.clicked.connect(self.on_dodaj_szafke)
        btns.addWidget(self.DodajSzBt)
        
        top.addLayout(btns)

        root.addLayout(top)
        # Mała tabela filtrów nad główną tabelą
        self.TabelaSzafekFilters = QtWidgets.QTableWidget(tab)
        self.TabelaSzafekFilters.setObjectName("TabelaSzafekFilters")
        self.TabelaSzafekFilters.setColumnCount(14)
        self.TabelaSzafekFilters.setRowCount(1)
        self.TabelaSzafekFilters.verticalHeader().setVisible(False)
        self.TabelaSzafekFilters.horizontalHeader().setVisible(False)
        self.TabelaSzafekFilters.setSelectionMode(QtWidgets.QAbstractItemView.SelectionMode.NoSelection)
        self.TabelaSzafekFilters.setFixedHeight(34)

        v = QtWidgets.QVBoxLayout()
        v.setSpacing(0)
        v.setContentsMargins(0,0,0,0)
        v.addWidget(self.TabelaSzafekFilters)
        v.addWidget(self.TabelaSzafek, 1)

        # Ukrycie numeracji wierszy dla wyrównania z filtrami
        try:
            self.TabelaSzafek.verticalHeader().setVisible(False)
        except Exception:
            pass

        root.addLayout(v, 1)

        # Udostępnienie filter_table dla renderera
        self.TabelaSzafek.filter_table = self.TabelaSzafekFilters

        # Synchronizacja scrollowania i szerokości kolumn
        self.TabelaSzafek.horizontalScrollBar().valueChanged.connect(self.TabelaSzafekFilters.horizontalScrollBar().setValue)
        self.TabelaSzafekFilters.horizontalScrollBar().valueChanged.connect(self.TabelaSzafek.horizontalScrollBar().setValue)
        self.TabelaSzafek.horizontalHeader().sectionResized.connect(lambda idx, old, new: self.TabelaSzafekFilters.setColumnWidth(idx, new))
        # Inicjalizacja szerokości kolumn
        for i in range(14):
            self.TabelaSzafekFilters.setColumnWidth(i, self.TabelaSzafek.columnWidth(i))

        # Checkbox dla filtra zwolnionych pracowników
        filter_row = QtWidgets.QHBoxLayout()
        filter_row.setContentsMargins(0, 8, 0, 0)
        
        self.ShowDismissedCh = QtWidgets.QCheckBox("Pokaż szafki zwolnionych pracowników")
        self.ShowDismissedCh.setChecked(False)
        filter_row.addWidget(self.ShowDismissedCh)
        filter_row.addStretch(1)
        
        root.addLayout(filter_row)

    def _layout_pracownicy_tab(self):
        tab = self.PracownicyTab
        if tab.layout() is not None:
            return
        root = QtWidgets.QVBoxLayout(tab)
        root.setContentsMargins(10, 10, 10, 10)
        root.setSpacing(8)

        top = QtWidgets.QHBoxLayout()
        top.setSpacing(8)

        top.addWidget(self.groupBox_2)
        top.addStretch(1)

        btns = QtWidgets.QHBoxLayout()
        btns.setSpacing(8)
        # checkbox tylko zatrudnieni
        self.OnlyActiveCh = QtWidgets.QCheckBox("Zatrudnieni")
        self.OnlyActiveCh.setChecked(True)
        self.OnlyActiveCh.setToolTip("Pokaż tylko aktualnie zatrudnionych pracowników")
        self.OnlyActiveCh.stateChanged.connect(self._on_toggle_only_active)
        btns.addWidget(self.OnlyActiveCh)
        # Ukryj przycisk PrzydzielSzPr - przydział tylko przez zakładkę Szafki
        self.PrzydzielSzPr.hide()
        btns.addWidget(self.UsunPr)
        btns.addWidget(self.NowiPracownicyBt)
        top.addLayout(btns)

        root.addLayout(top)

        # Lewa strona: filtr + tabela pracowników
        self.TabelaPracownikowFilters = QtWidgets.QTableWidget(tab)
        self.TabelaPracownikowFilters.setObjectName("TabelaPracownikowFilters")
        self.TabelaPracownikowFilters.setColumnCount(9)
        self.TabelaPracownikowFilters.setRowCount(1)
        self.TabelaPracownikowFilters.verticalHeader().setVisible(False)
        self.TabelaPracownikowFilters.horizontalHeader().setVisible(False)
        self.TabelaPracownikowFilters.setSelectionMode(QtWidgets.QAbstractItemView.SelectionMode.NoSelection)
        self.TabelaPracownikowFilters.setFixedHeight(34)

        left = QtWidgets.QWidget()
        left_layout = QtWidgets.QVBoxLayout(left)
        left_layout.setContentsMargins(0,0,0,0)
        left_layout.setSpacing(0)
        left_layout.addWidget(self.TabelaPracownikowFilters)
        left_layout.addWidget(self.TabelaPracownikow)

        split = QtWidgets.QSplitter(Qt.Orientation.Horizontal, tab)
        split.addWidget(left)
        split.addWidget(self.PanelSzafekPr)
        split.setStretchFactor(0, 9)
        split.setStretchFactor(1, 1)

        root.addWidget(split, 1)

        # Ustawienie początkowego podziału (prawy panel ~15%)
        try:
            from PyQt6.QtCore import QTimer
            def _set_split_sizes():
                try:
                    total = split.size().width() or self.width() or 800
                    # Panel ma stałą szerokość 480px
                    try:
                        if hasattr(self.PanelSzafekPr, 'setMinimumWidth'):
                            self.PanelSzafekPr.setMinimumWidth(0)
                    except Exception:
                        pass
                    # Podział 85%/15%
                    split.setSizes([int(total * 0.85), int(total * 0.15)])
                    # Powtórne ustawienie po krótkim opóźnieniu
                    try:
                        QTimer.singleShot(120, lambda: split.setSizes([int((split.size().width() or self.width() or 800) * 0.85), int((split.size().width() or self.width() or 800) * 0.15)]))
                    except Exception:
                        pass
                except Exception:
                    try:
                        split.setSizes([17, 3])
                    except Exception:
                        pass
            QTimer.singleShot(0, _set_split_sizes)
        except Exception:
            pass

        # Udostępnienie filter_table dla renderera
        self.TabelaPracownikow.filter_table = self.TabelaPracownikowFilters

        # Synchronizacja scrollowania i szerokości kolumn
        self.TabelaPracownikow.horizontalScrollBar().valueChanged.connect(self.TabelaPracownikowFilters.horizontalScrollBar().setValue)
        self.TabelaPracownikowFilters.horizontalScrollBar().valueChanged.connect(self.TabelaPracownikow.horizontalScrollBar().setValue)
        self.TabelaPracownikow.horizontalHeader().sectionResized.connect(lambda idx, old, new: self.TabelaPracownikowFilters.setColumnWidth(idx, new))
        # Inicjalizacja szerokości kolumn
        for i in range(9):
            self.TabelaPracownikowFilters.setColumnWidth(i, self.TabelaPracownikow.columnWidth(i))

        # Ukrycie numeracji wierszy dla wyrównania
        try:
            self.TabelaPracownikow.verticalHeader().setVisible(False)
        except Exception:
            pass

    # ---------- SYGNAŁY ----------
    def _connect_signals(self):
        # Szafki
        self.EdytujSz.clicked.connect(self.on_edytuj_szafke)
        self.ZwolnijSz.clicked.connect(self.on_zwolnij_szafke)
        self.PrzydzielSz.clicked.connect(self.on_przydziel_szafke)
        # Menu kontekstowe dla tabeli szafek
        self.TabelaSzafek.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.TabelaSzafek.customContextMenuRequested.connect(self._show_locker_context_menu)

        # Sygnał checkboxa zwolnionych
        self.ShowDismissedCh.toggled.connect(self._on_show_dismissed_toggled)

        # Stare filtry usunięte - używamy filtrów w tabeli

        # Pracownicy
        self.PrzydzielSzPr.clicked.connect(self.on_przydziel_szafke_pracownikowi)
        self.UsunPr.clicked.connect(self._zwolnij_wszystkie_szafki_pracownika)

        self.TabelaPracownikow.itemSelectionChanged.connect(self._odswiez_panel_szafek_pracownika)

    def _on_locker_double_click(self, index):
        """Podwójne kliknięcie: przydział jeśli pusta, edycja jeśli zajęta"""
        row = index.row()
        if row < 0:
            return
        
        # Pobranie ID szafki z kolumny 0
        id_item = self.TabelaSzafek.item(row, 0)
        if not id_item:
            return
        
        try:
            locker_id = int(id_item.text())
        except (ValueError, AttributeError):
            return
        
        # Szukanie szafki w cache
        target_locker = None
        for locker in getattr(self, '_wszystkie_szafki', []):
            if locker and locker[0] == locker_id:
                target_locker = locker
                break
        
        if target_locker is None:
            return
        
        # Kolumna 5 to kod pracownika
        kod_pracownika = str(target_locker[5] or '').strip() if len(target_locker) > 5 else ''
        
        if not kod_pracownika:
            # Pusta szafka - dialog przydzielania
            self.on_przydziel_szafke()
        else:
            # Zajęta szafka - dialog edycji
            self.on_edytuj_szafke()

    # ---------- PRZEŁADOWANIE ----------
    def reload_all(self):
        from session import get_current_user
        u = get_current_user()
        self._wszystkie_szafki = self.lockers.get_all(current_user=u)
        self._szafki_by_kod = self.lockers.map_by_employee_code(self._wszystkie_szafki)

        self._reload_miejsca_combobox()
        self.filtruj_tabele_szafek()
        self.filtruj_tabele_pracownikow()
        self._odswiez_panel_szafek_pracownika()
        self.sprawdz_nowych_pracownikow(pokaz_zawsze=False)

    def _reload_miejsca_combobox(self):
        miejsca = sorted(set(r[1] for r in self._wszystkie_szafki if r and len(r) >= 2 and r[1]))
        self.MiejsceCB.blockSignals(True)
        self.MiejsceCB.clear()
        self.MiejsceCB.addItem("")
        self.MiejsceCB.addItems(miejsca)
        self.MiejsceCB.setEnabled(True)
        self.MiejsceCB.blockSignals(False)

    # ---------- POMOCNICZE ----------
    def _selected_employee_code(self) -> str:
        row = self.TabelaPracownikow.currentRow()
        if row < 0:
            return ""
        item = self.TabelaPracownikow.item(row, 0)
        return (item.text() if item else "").strip()

    # ---------- PANEL SZAFEK PRACOWNIKA ----------
    def _odswiez_panel_szafek_pracownika(self):
        kod = self._selected_employee_code()
        if not kod:
            self.LblPracownikInfo.setText("Wybierz pracownika z tabeli po lewej.")
            self.TabelaSzafekPr.setRowCount(0)
            return

        nazw = self.TabelaPracownikow.item(self.TabelaPracownikow.currentRow(), 1).text()
        im = self.TabelaPracownikow.item(self.TabelaPracownikow.currentRow(), 2).text()

        szafki = self._szafki_by_kod.get(kod, [])
        self.LblPracownikInfo.setText(f"<b>{kod}</b> — {nazw} {im}<br>Przypisane szafki: <b>{len(szafki)}</b>")
        render_employee_lockers_panel(self.TabelaSzafekPr, szafki)

    def _zwolnij_szafke_z_panelu(self):
        r = self.TabelaSzafekPr.currentRow()
        if r < 0:
            QMessageBox.warning(self, "Brak zaznaczenia", "Zaznacz szafkę w panelu po prawej.")
            return
        id_item = self.TabelaSzafekPr.item(r, 3)
        if not id_item or not id_item.text().strip().isdigit():
            return
        locker_id = int(id_item.text().strip())

        potw = QMessageBox.question(
            self, "Potwierdź",
            f"Zwolnić szafkę (ID={locker_id}) z tego pracownika?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        if potw != QMessageBox.StandardButton.Yes:
            return

        self.lockers.release_locker_by_id(locker_id, rows=self._wszystkie_szafki)
        self.reload_all()

    def _zwolnij_wszystkie_szafki_pracownika(self):
        kod = self._selected_employee_code()
        if not kod:
            QMessageBox.warning(self, "Brak pracownika", "Wybierz pracownika z tabeli po lewej.")
            return

        count = len(self._szafki_by_kod.get(kod, []))
        if count == 0:
            QMessageBox.information(self, "Brak szafek", "Ten pracownik nie ma przypisanych szafek.")
            return

        potw = QMessageBox.question(
            self, "Potwierdź",
            f"Zwolnić WSZYSTKIE szafki pracownika {kod}? (liczba: {count})",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        if potw != QMessageBox.StandardButton.Yes:
            return

        self.lockers.release_all_for_employee(kod, rows=self._wszystkie_szafki)
        self.reload_all()

    def _skocz_do_szafki_w_zakladce(self):
        r = self.TabelaSzafekPr.currentRow()
        if r < 0:
            return
        id_item = self.TabelaSzafekPr.item(r, 3)
        if not id_item:
            return
        id_sz = id_item.text().strip()
        if not id_sz:
            return

        self.SzafkiWidget.setCurrentIndex(0)
        for i in range(self.TabelaSzafek.rowCount()):
            it = self.TabelaSzafek.item(i, 0)
            if it and it.text().strip() == id_sz:
                self.TabelaSzafek.selectRow(i)
                self.TabelaSzafek.scrollToItem(self.TabelaSzafek.item(i, 1))
                break

    def _init_account_menu(self):
        menu = self.menuBar().addMenu("Konto")
        change_pw_action = QtGui.QAction("Zmień hasło", self)
        change_pw_action.triggered.connect(self._change_password)
        menu.addAction(change_pw_action)
        logout_action = QtGui.QAction("Wyloguj", self)
        logout_action.triggered.connect(self._logout_and_relogin)
        menu.addAction(logout_action)

    def _change_password(self):
        u = get_current_user()
        if not u:
            QMessageBox.warning(self, "Brak użytkownika", "Nie ma zalogowanego użytkownika.")
            return
        dlg = ChangePasswordDialog(u['login'], self, require_old=True)
        if dlg.exec() == QDialog.DialogCode.Accepted:
            QMessageBox.information(self, "Sukces", "Hasło zmienione.")
            # Aktualizacja statusu użytkownika
            self._update_user_status()

    def _update_user_status(self):
        u = get_current_user()
        if not u:
            self.statusBar().showMessage("Brak zalogowanego użytkownika")
        else:
            self.statusBar().showMessage(f"Zalogowany: {u.get('login')} ({u.get('role')}, dzial={u.get('dzial')})")

    def _logout_and_relogin(self):
        potw = QMessageBox.question(self, "Wyloguj", "Czy chcesz się wylogować?",
                                    QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        if potw != QMessageBox.StandardButton.Yes:
            return
        clear_current_user()
        # Zamknij bieżące główne okno
        self.close()

        # Prompt logowania
        dlg = LoginDialog(self)
        if dlg.exec() != QDialog.DialogCode.Accepted:
            import sys
            sys.exit(0)
            return

        # Po pomyślnym ponownym logowaniu: otwórz nowe główne okno dla nowej sesji
        new_win = OknoGlowne()
        new_win.show()
        # Nic więcej do zrobienia; wywołujący będzie kontynuował z nowym oknem

    def _ensure_admin_tabs(self):
        if self._admin_tabs_created:
            return
        u = get_current_user()
        if not u or u.get('role') != 'admin':
            return
        # Utwórz zakładkę 'Administracja' z podzakładkami 'Użytkownicy' i 'Historia'
        admin_tab = QtWidgets.QWidget()
        admin_layout = QtWidgets.QVBoxLayout(admin_tab)
        users_panel = UsersPanel(admin_tab)
        admin_layout.addWidget(users_panel)
        from ui.admin.history_panel import HistoryPanel
        history_panel = HistoryPanel(admin_tab)
        admin_layout.addWidget(history_panel)
        idx = self.SzafkiWidget.addTab(admin_tab, "Administracja")
        self.SzafkiWidget.setCurrentIndex(idx)
        self._admin_tabs_created = True

    # ---------- FILTRY I RENDEROWANIE ----------
    def filtruj_tabele_szafek(self):
        # Tylko filtry z tabeli filtrów
        from session import get_current_user
        u = get_current_user()
        is_admin = bool(u and u.get('role') == 'admin')
        visible_cols = 14 + (1 if is_admin else 0)

        # Filtrowanie tylko widocznych kolumn
        active_filters = {c: v for c, v in self._column_filters.items() if c < visible_cols}

        # Słownik pracowników do sprawdzenia zwolnień
        def s(x): return "" if x is None else str(x)
        employees_data = {}
        df = DF_PRACOWNICY.copy().reset_index()
        for _, r in df.iterrows():
            kod = s(r.get("Kod")).strip()
            if kod:
                employees_data[kod] = {
                    'Data_zwolnienia': s(r.get("Data_zwolnienia")).strip()
                }

        filtered = []
        for r in self._wszystkie_szafki:
            if not r or len(r) < 14:
                continue

            # Filtrowanie kolumn z tabeli filtrów
            skip = False
            for col, filt in active_filters.items():
                if not filt:
                    continue
                val = str(r[col] or "").lower() if col < len(r) else ""
                if filt not in val:
                    skip = True
                    break
            if skip:
                continue

            # Filtr szafek zwolnionych pracowników
            if self._show_dismissed_lockers and len(r) > 5:
                employee_kod = str(r[5] or "").strip()
                # Jeśli checkbox zaznaczony - tylko szafki zwolnionych
                if not employee_kod:
                    continue
                if employee_kod not in employees_data:
                    continue
                emp_data = employees_data[employee_kod]
                # Jeśli brak daty zwolnienia - pomiń
                if not emp_data.get('Data_zwolnienia', '').strip():
                    continue

            filtered.append(r)

        render_lockers_table(self.TabelaSzafek, filtered, filter_callback=self._on_locker_filter_changed, filter_values=active_filters, current_user=u, employees_data=employees_data)

    def _on_locker_filter_changed(self, col: int, text: str):
        # Zapis filtra (lowercase, trimmed)
        txt = (text or "").strip().lower()
        if txt:
            self._column_filters[col] = txt
        elif col in self._column_filters:
            del self._column_filters[col]
        # Ponowne filtrowanie
        self.filtruj_tabele_szafek()

    def _on_show_dismissed_toggled(self, checked: bool):
        """Obsługa przełączenia checkboxa zwolnionych"""
        self._show_dismissed_lockers = checked
        self.filtruj_tabele_szafek()

    def on_kod_changed(self, ui):
        kod = ui.KodPracDP.text().strip()
        if not kod or kod not in DF_PRACOWNICY.index:
            if hasattr(ui, "NazwiskoDP"): ui.NazwiskoDP.setText("")
            if hasattr(ui, "ImieDP"): ui.ImieDP.setText("")
            if hasattr(ui, "PlecDP"): ui.PlecDP.setText("")
            if hasattr(ui, "DzialDP"): ui.DzialDP.setText("")
            if hasattr(ui, "StanowiskoDP"): ui.StanowiskoDP.setText("")
            if hasattr(ui, "ZmianaDP"): ui.ZmianaDP.setText("")
            if hasattr(ui, "MiejsceDP"): ui.MiejsceDP.setText("")
            return

        p = DF_PRACOWNICY.loc[kod]
        if hasattr(ui, "NazwiskoDP"): ui.NazwiskoDP.setText(str(p.get("Nazwisko","")))
        if hasattr(ui, "ImieDP"): ui.ImieDP.setText(str(p.get("Imię","")))
        if hasattr(ui, "PlecDP"): ui.PlecDP.setText(str(p.get("Płeć","")))
        if hasattr(ui, "DzialDP"): ui.DzialDP.setText(str(p.get("Dział","")))
        if hasattr(ui, "StanowiskoDP"): ui.StanowiskoDP.setText(str(p.get("Stanowisko","")))
        if hasattr(ui, "ZmianaDP"): ui.ZmianaDP.setText(str(p.get("Zmiana","")))
        if hasattr(ui, "MiejsceDP"): 
            ui.MiejsceDP.setText("")
        
    def filtruj_tabele_pracownikow(self):
        # Budowanie pełnej listy
        df = DF_PRACOWNICY.copy().reset_index()

        def s(x): return "" if x is None else str(x)

        lista = []
        for _, r in df.iterrows():
            kod  = s(r.get("Kod")).strip()
            nazw = s(r.get("Nazwisko")).strip()
            im   = s(r.get("Imię")).strip()
            dz   = s(r.get("Dział")).strip()
            stan = s(r.get("Stanowisko")).strip()
            plec = s(r.get("Płeć")).strip()
            dzatr = s(r.get("Data_zatrudnienia")).strip()
            dzw   = s(r.get("Data_zwolnienia")).strip()

            sz_list = self._szafki_by_kod.get(kod, [])
            lista.append({
                "Kod": kod,
                "Nazwisko": nazw,
                "Imię": im,
                "Płeć": plec,
                "Dział": dz,
                "Stanowisko": stan,
                "Data_zatrudnienia": dzatr,
                "Data_zwolnienia": dzw,
                "Szafki_count": len(sz_list),
            })

        # Filtr 'Zatrudnieni' przed filtrami kolumn
        if getattr(self, '_emp_only_active', True):
            lista = [e for e in lista if ((e.get('Data_zwolnienia') or '').strip() == '')]

        # Filtrowanie kolumn z tabeli filtrów
        filtered = []
        for e in lista:
            skip = False
            # Mapowanie: 0=Kod,1=Nazwisko,2=Imię,3=Płeć,4=Dział,5=Stanowisko,6=Data zatr.,7=Data zwol.,8=Szafki
            values = [
                e.get("Kod", "").lower(), e.get("Nazwisko", "").lower(), e.get("Imię", "").lower(),
                e.get("Płeć", "").lower(), e.get("Dział", "").lower(), e.get("Stanowisko", "").lower(),
                e.get("Data_zatrudnienia", "").lower(), e.get("Data_zwolnienia", "").lower(),
                str(e.get("Szafki_count", 0)).lower()
            ]
            for col, filt in self._emp_column_filters.items():
                if not filt:
                    continue
                val = values[col] if col < len(values) else ""
                if filt not in val:
                    skip = True
                    break
            if skip:
                continue
            filtered.append(e)

        render_employees_table(self.TabelaPracownikow, filtered, filter_callback=self._on_employee_filter_changed, filter_values=self._emp_column_filters)

    def _on_employee_filter_changed(self, col: int, text: str):
        txt = (text or "").strip().lower()
        if txt:
            self._emp_column_filters[col] = txt
        elif col in self._emp_column_filters:
            del self._emp_column_filters[col]
        self.filtruj_tabele_pracownikow()

    def _on_toggle_only_active(self, state:int):
        try:
            from PyQt6.QtCore import Qt
            self._emp_only_active = bool(state == Qt.Checked)
        except Exception:
            self._emp_only_active = bool(state)
        self.filtruj_tabele_pracownikow()

    # ---------- NOWI PRACOWNICY ----------
    def sprawdz_nowych_pracownikow(self, *, pokaz_zawsze=False):
        new_codes = self.employees.find_new_active_without_lockers(self._wszystkie_szafki)

        if not new_codes:
            if pokaz_zawsze:
                QMessageBox.information(self, "Nowi pracownicy", "Nie znaleziono nowych aktywnych pracowników bez szafki.")
            return

        df_new = DF_PRACOWNICY.loc[new_codes].reset_index()

        # Data graniczna na podstawie ostatniego zdarzenia użytkownika
        try:
            from history import get_events
            from session import get_current_user
            from datetime import datetime, date
            u = get_current_user()
            last_dt = None
            if u and u.get('login'):
                events = get_events(1000)
                for ev in events:
                    if ev.get('performed_by') == u.get('login'):
                        ts = ev.get('ts')
                        try:
                            last_dt = datetime.fromisoformat(ts.replace('Z', ''))
                        except Exception:
                            try:
                                last_dt = datetime.fromisoformat(ts)
                            except Exception:
                                last_dt = None
                        if last_dt:
                            break
            # Domyślna data graniczna: 2025-11-30
            if last_dt is None:
                cutoff = date(2025, 11, 30)
            else:
                cutoff = last_dt.date()
            # Filtrowanie pracowników zatrudnionych po dacie granicznej
            def _parse_date(s):
                if not s or str(s).strip() == "":
                    return None
                s = str(s).strip()
                for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d-%m-%Y"):
                    try:
                        return datetime.strptime(s, fmt).date()
                    except Exception:
                        pass
                # Ostatnia próba z pandas
                try:
                    import pandas as pd
                    dt = pd.to_datetime(s, dayfirst=True, errors='coerce')
                    if pd.isna(dt):
                        return None
                    return dt.date()
                except Exception:
                    return None

            df_new['__hire_date'] = df_new.get('Data_zatrudnienia', '').apply(_parse_date)
            df_new = df_new[df_new['__hire_date'].apply(lambda d: d is not None and d > cutoff)]
            df_new = df_new.drop(columns=['__hire_date'])
        except Exception:
            # W razie błędu - oryginalny zestaw
            pass

        if df_new.empty:
            if pokaz_zawsze:
                QMessageBox.information(self, "Nowi pracownicy", "Nie znaleziono nowych aktywnych pracowników spełniających kryteria.")
            return

        dlg = NowiPracownicyDialog(df_new, parent=self)
        dlg.exec()

    def on_przydziel_szafke(self):
        try:
            # Przydzielanie szafki wybranej w tabeli
            row = self.TabelaSzafek.currentRow()
            if row < 0:
                QMessageBox.warning(self, "Brak zaznaczenia", "Zaznacz szafkę.")
                return
            id_item = self.TabelaSzafek.item(row, 0)
            if not id_item or not id_item.text().strip().isdigit():
                QMessageBox.warning(self, "Błąd", "Nie mogę odczytać ID szafki.")
                return
            locker_id = int(id_item.text().strip())

            # Szukanie szafki w cache
            target = next((r for r in self._wszystkie_szafki if r and r[0] == locker_id), None)
            if not target:
                QMessageBox.warning(self, "Błąd", "Nie znaleziono szafki w pamięci aplikacji.")
                return

            # Sprawdzenie czy szafka jest zajęta lub nieczynna
            kod_pracownika = str(target[5] or '').strip() if len(target) > 5 else ''
            status = str(target[12] or '').strip() if len(target) > 12 else ''
            
            if kod_pracownika:
                QMessageBox.warning(self, "Szafka zajęta", "Ta szafka jest już przydzielona pracownikowi. Zwolnij ją najpierw.")
                return
            
            if status == "Nieczynna":
                QMessageBox.warning(self, "Szafka nieczynna", "Ta szafka jest w stanie nieczynnym. Zmień status na 'Wolna'.")
                return

            # Sprawdzenie uprawnień: zwykli użytkownicy tylko w swoim dziale
            try:
                from session import get_current_user
                u = get_current_user()
                if u and u.get('role') != 'admin':
                    allowed = (u.get('dzial') or '').strip()
                    physical = str((target[14] if len(target) > 14 else '') or '').strip()
                    if allowed and physical and allowed != physical:
                        QMessageBox.warning(self, "Brak uprawnień", "Możesz przydzielać tylko szafki w swoim dziale.")
                        return
            except Exception:
                pass

            # Dialog przydzielania szafki
            from ui.dialogs.przydziel_szafke_dialog import PrzydzielSzafkeDialog
            
            locker_info = {
                'Miejsce': target[1],
                'Nr_szafki': target[2],
                'Płeć_szatni': target[4],
            }
            
            dlg = PrzydzielSzafkeDialog(locker_info, self)
            if dlg.exec() != QDialog.DialogCode.Accepted:
                return
            
            selected_emp = dlg.get_selected_employee()
            if not selected_emp:
                QMessageBox.warning(self, "Błąd", "Nie wybrano pracownika.")
                return
            
            kod = selected_emp.get('Kod', '').strip()
            if not kod:
                QMessageBox.warning(self, "Błąd", "Nie można odczytać kodu pracownika.")
                return

            # Walidacja statusu pracownika
            if kod not in DF_PRACOWNICY.index:
                QMessageBox.warning(self, "Błąd", f"Pracownik o kodzie {kod} nie istnieje.")
                return
                
            p = DF_PRACOWNICY.loc[kod]
            dz_zw = str(p.get("Data_zwolnienia", "") or "").strip()
            if dz_zw:
                QMessageBox.warning(self, "Zwolniony", "Tego pracownika nie można przypisać (zwolniony).")
                return

            # Sprawdzenie zgodności płci
            locker_plec = str(target[4] or "").strip()
            emp_plec = str(p.get("Płeć", "") or "").strip()
            
            # Mapowanie płci pracownika na płeć szatni
            emp_to_locker = {'Mężczyzna': 'Męska', 'Kobieta': 'Damska'}
            emp_locker_plec = emp_to_locker.get(emp_plec, emp_plec)
            
            if locker_plec and locker_plec != 'Neutralna' and emp_locker_plec and emp_locker_plec != locker_plec:
                QMessageBox.warning(self, "Niepasująca płeć", "Ta szatnia jest przypisana do innej płci i nie można jej przydzielić temu pracownikowi.")
                return

            # Przypisanie szafki
            ok = self.assignment.assign_locker_to_employee(
                locker_id,
                kod_pracownika=kod,
                nazwisko=str(p.get('Nazwisko','') or ''),
                imie=str(p.get('Imię','') or ''),
                dzial=str(p.get('Dział','') or ''),
                stanowisko=str(p.get('Stanowisko','') or ''),
                plec=str(p.get('Płeć','') or ''),
                status="Zajęta",
            )

            if not ok:
                QMessageBox.warning(self, "Błąd", "Nie udało się przypisać szafki.")
                return

            QMessageBox.information(self, "Sukces", "Przypisano szafkę.")
            self.reload_all()
        except Exception as e:
            import traceback
            tb = traceback.format_exc()
            print("Error in on_przydziel_szafke:", tb)
            try:
                QMessageBox.critical(self, "Błąd", f"Wystąpił błąd podczas przydziału szafki:\n{str(e)}")
            except Exception:
                pass

    # ---------- MENU KONTEKSTOWE I EKSPORT ----------
    def _show_locker_context_menu(self, pos):
        """Menu kontekstowe dla tabeli szafek"""
        row = self.TabelaSzafek.rowAt(pos.y())
        if row < 0:
            return
        
        # Pobranie ID szafki z kolumny 0
        id_item = self.TabelaSzafek.item(row, 0)
        if not id_item:
            return
        
        try:
            locker_id = int(id_item.text())
        except (ValueError, AttributeError):
            return
        
        # Szukanie szafki w cache
        target_locker = None
        for locker in self._wszystkie_szafki:
            if locker and locker[0] == locker_id:
                target_locker = locker
                break
        
        if target_locker is None:
            return
        
        # Sprawdzenie czy szafka jest zajęta
        kod_pracownika = str(target_locker[5] or '').strip() if len(target_locker) > 5 else ''
        is_occupied = bool(kod_pracownika)
        
        # Status szafki (kolumna 12)
        status = str(target_locker[12] or '').strip() if len(target_locker) > 12 else ''
        
        # Tworzenie menu kontekstowego
        menu = QtWidgets.QMenu(self)
        
        if is_occupied:
            # Opcje dla zajętej szafki
            release_action = menu.addAction("Zwolnij szafkę")
            release_action.triggered.connect(lambda: self._context_release_locker(locker_id))
            edit_action = menu.addAction("Edytuj")
            edit_action.triggered.connect(lambda: self._context_edit_locker(locker_id))
        else:
            # Opcje dla pustej szafki
            assign_action = menu.addAction("Przydziel szafkę")
            assign_action.triggered.connect(lambda: self._context_assign_locker(locker_id))
            edit_action = menu.addAction("Edytuj")
            edit_action.triggered.connect(lambda: self._context_edit_locker(locker_id))
            
            # Przełączanie statusu dla pustych szafek
            if status == "Wolna":
                toggle_status_action = menu.addAction("Zmień status na: Nieczynna")
                toggle_status_action.triggered.connect(lambda: self._context_toggle_status(locker_id, "Nieczynna"))
            elif status == "Nieczynna":
                toggle_status_action = menu.addAction("Zmień status na: Wolna")
                toggle_status_action.triggered.connect(lambda: self._context_toggle_status(locker_id, "Wolna"))
        
        # Wspólne opcje
        delete_action = menu.addAction("Usuń szafkę")
        delete_action.triggered.connect(lambda: self._context_delete_locker(locker_id))
        
        menu.exec(self.TabelaSzafek.mapToGlobal(pos))
    
    def _context_release_locker(self, locker_id):
        """Zwolnienie szafki z menu kontekstowego"""
        self.TabelaSzafek.selectRow(0)  # Resetuj zaznaczenie, aby uniknąć problemów z UI
        self.lockers.release_locker_by_id(locker_id, rows=self._wszystkie_szafki)
        self.reload_all()
    
    def _context_assign_locker(self, locker_id):
        """Przydzielenie szafki z menu kontekstowego"""
        # Zaznacz wiersz i wywołaj handler przydziału
        for i in range(self.TabelaSzafek.rowCount()):
            item = self.TabelaSzafek.item(i, 0)
            if item and item.text().strip() == str(locker_id):
                self.TabelaSzafek.selectRow(i)
                break
        self.on_przydziel_szafke()
    
    def _context_edit_locker(self, locker_id):
        """Edycja szafki z menu kontekstowego"""
        # Zaznacz wiersz i wywołaj handler edycji
        for i in range(self.TabelaSzafek.rowCount()):
            item = self.TabelaSzafek.item(i, 0)
            if item and item.text().strip() == str(locker_id):
                self.TabelaSzafek.selectRow(i)
                break
        self.on_edytuj_szafke()
    
    def _context_toggle_status(self, locker_id, new_status):
        """Zmiana statusu szafki z menu kontekstowego"""
        try:
            # Szukanie szafki w cache
            target_locker = None
            for locker in self._wszystkie_szafki:
                if locker and locker[0] == locker_id:
                    target_locker = locker
                    break
            
            if target_locker is None:
                QMessageBox.warning(self, "Błąd", "Nie znaleziono szafki w pamięci cache.")
                return
            
            # Stary status przed zmianą
            old_status = target_locker[12] if len(target_locker) > 12 else ""
            
            # Pobranie danych szafki
            edytuj_szafke(
                id_szafki=target_locker[0],
                miejsce=target_locker[1],
                nr_szafki=target_locker[2],
                nr_zamka=target_locker[3],
                plec_szatni=target_locker[4],
                kod_pracownika=target_locker[5],
                dzial=target_locker[8],
                status=new_status,
                komentarz=target_locker[13] if len(target_locker) > 13 else ""
            )
            
            # Log zdarzenia
            from history import log_event
            log_event(
                event_type="edit",
                locker_id=locker_id,
                before={'status': old_status},
                after={'status': new_status}
            )
            
            QMessageBox.information(self, "Sukces", f"Status szafki zmieniony na: {new_status}")
            self.reload_all()
        except Exception as e:
            QMessageBox.critical(self, "Błąd", f"Nie udało się zmienić statusu szafki:\n{str(e)}")
    
    def _context_delete_locker(self, locker_id):
        """Usunięcie szafki z menu kontekstowego"""
        potw = QMessageBox.question(
            self, "Potwierdź",
            f"Czy na pewno usunąć szafkę (ID={locker_id})?\nTej operacji nie można cofnąć.",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        if potw != QMessageBox.StandardButton.Yes:
            return
        
        try:
            usun_szafke(locker_id)
            from history import log_event
            log_event("delete", locker_data={'id': locker_id})
            QMessageBox.information(self, "Sukces", f"Szafka {locker_id} usunięta.")
            self.reload_all()
        except Exception as e:
            QMessageBox.critical(self, "Błąd", f"Nie udało się usunąć szafki:\n{str(e)}")
    
    def _init_export_menu(self):
        """Inicjalizacja menu Eksport"""
        menu = self.menuBar().addMenu("Eksport")
        
        csv_action = QtGui.QAction("Do CSV", self)
        csv_action.triggered.connect(self._export_lockers_csv)
        menu.addAction(csv_action)
        
        excel_action = QtGui.QAction("Do Excel", self)
        excel_action.triggered.connect(self._export_lockers_excel)
        menu.addAction(excel_action)
    
    def _export_lockers_csv(self):
        """Eksport do CSV (szafki lub pracownicy)"""
        import csv
        from PyQt6.QtWidgets import QFileDialog
        
        # Określ, którą tabelę eksportować na podstawie aktywnej zakładki
        current_tab_index = self.SzafkiWidget.currentIndex()
        if current_tab_index == 1:  # Pracownicy tab
            table = self.TabelaPracownikow
            export_type = "pracowników"
            default_name = "pracownicy.csv"
        else:  # Szafki tab (default)
            table = self.TabelaSzafek
            export_type = "szafki"
            default_name = "szafki.csv"
        
        file_path, _ = QFileDialog.getSaveFileName(
            self, f"Eksportuj {export_type} do CSV", default_name, "CSV Files (*.csv)"
        )
        if not file_path:
            return
        
        try:
            # BOM dla Excela
            with open(file_path, 'w', newline='', encoding='utf-8-sig') as f:
                writer = csv.writer(f)
                
                # Nagłówki
                headers = []
                for col in range(table.columnCount()):
                    if not table.isColumnHidden(col):
                        header_item = table.horizontalHeaderItem(col)
                        headers.append(header_item.text() if header_item else f"Kolumna {col}")
                writer.writerow(headers)
                
                # Dane
                for row in range(table.rowCount()):
                    row_data = []
                    for col in range(table.columnCount()):
                        if not table.isColumnHidden(col):
                            item = table.item(row, col)
                            row_data.append(item.text() if item else "")
                    writer.writerow(row_data)
            
            # Log eksportu do historii
            self._log_export_event(export_type, file_path, table.rowCount())
            
            QMessageBox.information(self, "Sukces", f"Dane wyeksportowane do:\n{file_path}")
        except Exception as e:
            QMessageBox.critical(self, "Błąd", f"Nie udało się wyeksportować danych:\n{str(e)}")
    
    def _export_lockers_excel(self):
        """Eksport do Excel (szafki lub pracownicy)"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            QMessageBox.warning(
                self, "Brak biblioteki",
                "Biblioteka 'openpyxl' nie jest zainstalowana.\n"
                "Zainstaluj ją poleceniem: pip install openpyxl"
            )
            return
        
        from PyQt6.QtWidgets import QFileDialog
        
        # Wybór tabeli na podstawie aktywnej zakładki
        current_tab_index = self.SzafkiWidget.currentIndex()
        if current_tab_index == 1:  # Pracownicy tab
            table = self.TabelaPracownikow
            export_type = "pracowników"
            sheet_title = "Pracownicy"
            default_name = "pracownicy.xlsx"
        else:  # Szafki tab (domyślnie)
            table = self.TabelaSzafek
            export_type = "szafki"
            sheet_title = "Szafki"
            default_name = "szafki.xlsx"
        
        file_path, _ = QFileDialog.getSaveFileName(
            self, f"Eksportuj {export_type} do Excel", default_name, "Excel Files (*.xlsx)"
        )
        if not file_path:
            return
        
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = sheet_title
            
            # Nagłówki
            col_idx = 1
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for col in range(table.columnCount()):
                if not table.isColumnHidden(col):
                    header_item = table.horizontalHeaderItem(col)
                    header_text = header_item.text() if header_item else f"Kolumna {col}"
                    cell = ws.cell(row=1, column=col_idx, value=header_text)
                    cell.fill = header_fill
                    cell.font = header_font
                    col_idx += 1
            
            # Dane
            for row in range(table.rowCount()):
                col_idx = 1
                for col in range(table.columnCount()):
                    if not table.isColumnHidden(col):
                        item = table.item(row, col)
                        value = item.text() if item else ""
                        
                        cell = ws.cell(row=row + 2, column=col_idx, value=value)
                        
                        # Zwolnieni na czerwono z pogrubieniem
                        if item and item.foreground().color().name() == "#ff0000":
                            cell.font = Font(bold=True, color="FF0000")
                        
                        col_idx += 1
            
            # Autodopasowanie szerokości kolumn
            for col in range(1, col_idx):
                max_length = 0
                col_letter = openpyxl.utils.get_column_letter(col)
                for row in range(1, table.rowCount() + 2):
                    cell = ws[f'{col_letter}{row}']
                    if len(str(cell.value or "")) > max_length:
                        max_length = len(str(cell.value or ""))
                ws.column_dimensions[col_letter].width = min(max_length + 2, 50)
            
            wb.save(file_path)
            
            # Log eksportu do historii
            self._log_export_event(export_type, file_path, table.rowCount())
            
            QMessageBox.information(self, "Sukces", f"Dane wyeksportowane do:\n{file_path}")
        except Exception as e:
            QMessageBox.critical(self, "Błąd", f"Nie udało się wyeksportować danych:\n{str(e)}")

    def _log_export_event(self, export_type: str, file_path: str, row_count: int):
        """Log eksportu do historii (bez cofania)"""
        try:
            from history import log_event
            from session import get_current_user
            import os
            
            performed_by = (get_current_user() or {}).get('login')
            filename = os.path.basename(file_path)
            
            log_event(
                event_type="export",
                locker_id=0,  # Brak konkretnej szafki
                before=None,
                after={
                    'export_type': export_type,
                    'filename': filename,
                    'row_count': row_count
                },
                performed_by=performed_by
            )
        except Exception:
            pass  # Nie przerywaj eksportu, jeśli logowanie się nie powiedzie